# -*- coding: utf-8 -*-
# @Time     :2020/2/17 18:47
# @Author   :lei.qin
# @File     :do_excel.py
from openpyxl import load_workbook
from tools.read_config import ReadConfig
class Do_Excel():

    def __init__(self, file_name, sheet_name):
        self.file_name = file_name
        self.sheet_name = sheet_name

    def get_header(self):
        '''
            获取第一行标题行
        '''
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]
        header = []  # 存储标题行
        for j in range(1,sheet.max_column+1):
            header.append(sheet.cell(1,j).value)
        return header

    def get_data(self):
        '''mode控制是否执行所有用例，默认值为all。如果不等于all就不全部执行
            如果不等于all，进入分支判断
            mode的值只能如是all， 列表 这两种类型的参数
        '''
        #从配置文件读取model值
        mode = ReadConfig().read_config('../tools/case.config', 'MODE', 'mode')
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]
        header = self.get_header()  # 拿到header
        test_data = []

        for i in range(2,sheet.max_row+1):  # 1 2 3
            #sub_data = {}
            # for j in range(1, sheet.max_column+1):  # 1 2 3 4 5 6 7 8
            #     sub_data[header[j-1]] = sheet.cell(i, j).value
            # test_data.append(sub_data)
            #print(test_data)
            sub_data = {}
            sub_data['case_id'] = sheet.cell(i,1).value
            sub_data['module'] = sheet.cell(i,2).value
            sub_data['title'] = sheet.cell(i,3).value
            sub_data['url'] = sheet.cell(i,4).value
            sub_data['data'] = sheet.cell(i,5).value
            sub_data['except1'] = sheet.cell(i,6).value
            sub_data['except2'] = sheet.cell(i,7).value
            sub_data['http_method'] = sheet.cell(i,8).value
            test_data.append(sub_data)

         #根据button的值判断
        if mode  == 'all':
         final_data = test_data
        else:
            final_data = []
            for item in test_data:
                if item['case_id'] in eval(mode):
                    final_data.append(item)
        return final_data


if __name__ == '__main__':
    print(Do_Excel("../tools/test_data.xlsx", "oldyellowhistory").get_data())
